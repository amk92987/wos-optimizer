"""Quick script to read WOS Upgrades.xlsx and dump contents."""
import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\adam\IdeaProjects\wos-optimizer\Screenshots\WOS Upgrades.xlsx",
    data_only=True,
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"=== Sheet: {sheet_name} ({ws.max_row} rows x {ws.max_column} cols) ===")

    # Print just the data columns that have the upgrade tables (columns K-L and O-S)
    if sheet_name == "Blad1":
        print("\n--- Hero Gear Upgrade Table (Columns K-M: Level, Hero Stones, Hero Gear) ---")
        for row in ws.iter_rows(min_row=1, max_row=102, min_col=11, max_col=13, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                print(" | ".join(vals))

        print("\n--- Legendary Gear Upgrade Table (Columns O-S: Level, Mithril, Hero Gear, XP) ---")
        for row in ws.iter_rows(min_row=1, max_row=102, min_col=15, max_col=19, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                print(" | ".join(vals))

        print("\n--- Calculator Section (Columns A-B) ---")
        for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=2, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                print(" | ".join(vals))

        print("\n--- Gear Slots + Resources Section (Columns E-H) ---")
        for row in ws.iter_rows(min_row=1, max_row=40, min_col=5, max_col=9, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                print(" | ".join(vals))
    else:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            vals = [str(v) if v is not None else "" for v in row]
            if any(vals):
                print(" | ".join(vals))
    print()
